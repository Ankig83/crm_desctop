from __future__ import annotations

import json
import re
import sqlite3
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from crm_desktop.repositories import audit, clients, products, promotions
from crm_desktop.repositories.clients import CLIENT_TYPES
from crm_desktop.utils.bonus_ids import (
    missing_product_external_ids,
    normalize_product_external_ids_csv,
    parse_product_external_ids_csv,
)
from crm_desktop.utils.dates import format_dmY, iso, parse_dmY, parse_iso
from crm_desktop.utils.validation import normalize_inn


def _norm_header(s: Any) -> str:
    t = str(s or "").strip().lower().replace("ё", "е")
    return re.sub(r"\s+", " ", t)


@dataclass
class ImportReport:
    errors: list[str] = field(default_factory=list)
    clients_rows: int = 0
    products_rows: int = 0
    promotions_rows: int = 0


def _row_is_empty(row: tuple[Any, ...]) -> bool:
    return all(v is None or str(v).strip() == "" for v in row)


def _header_map(row: tuple[Any, ...]) -> dict[str, int]:
    m: dict[str, int] = {}
    for i, cell in enumerate(row):
        key = _norm_header(cell)
        if key:
            m[key] = i
    return m


def _cell(row: tuple[Any, ...], m: dict[str, int], *names: str) -> str:
    for n in names:
        k = _norm_header(n)
        if k in m:
            v = row[m[k]]
            if v is None:
                return ""
            return str(v).strip()
    return ""


def _raw_cell_by_index(row: tuple[Any, ...], idx: int) -> str:
    if idx < 0 or idx >= len(row):
        return ""
    v = row[idx]
    if v is None:
        return ""
    return str(v).strip()


def _float_cell(row: tuple[Any, ...], m: dict[str, int], *names: str) -> float | None:
    s = _cell(row, m, *names)
    if s == "":
        return None
    s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def _int_cell(row: tuple[Any, ...], m: dict[str, int], *names: str) -> int | None:
    v = _float_cell(row, m, *names)
    return int(v) if v is not None else None


def _find_header_row(
    rows: list[tuple[Any, ...]],
    required_headers: tuple[str, ...],
) -> tuple[int, tuple[Any, ...], dict[str, int]] | None:
    for idx, row in enumerate(rows):
        if _row_is_empty(row):
            continue
        hm = _header_map(row)
        if all(_norm_header(h) in hm for h in required_headers):
            return idx, row, hm
    return None


def _load_rows_with_openpyxl(path: Path) -> list[tuple[Any, ...]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    return [tuple(r) for r in ws.iter_rows(values_only=True)]


def _load_rows_with_calamine(path: Path) -> list[tuple[Any, ...]]:
    try:
        import pandas as pd
    except Exception as e:
        raise RuntimeError("Для fallback-импорта нужен pandas.") from e
    try:
        df = pd.read_excel(path, sheet_name=0, engine="calamine", header=None, dtype=str)
    except Exception as e:
        raise RuntimeError("Не удалось прочитать файл через calamine.") from e
    rows: list[tuple[Any, ...]] = []
    for rec in df.itertuples(index=False, name=None):
        cleaned: list[str] = []
        for v in rec:
            s = "" if v is None else str(v)
            if s.strip().lower() == "nan":
                s = ""
            cleaned.append(s)
        rows.append(tuple(cleaned))
    return rows


def _extract_matrix_rules_json(
    row: tuple[Any, ...],
    header_row: tuple[Any, ...],
    hm: dict[str, int],
) -> str:
    known = {
        _norm_header(h) for h in (
            "id товара", "id", "тип акции", "тип", "размер скидки", "скидка",
            "скидка %", "дата начала", "начало", "дата окончания", "окончание",
            "конец", "id товаров-бонусов (через запятую)", "id товаров бонуса",
            "id товаров-бонусов", "бонусные id", "id бонусных товаров",
            "id бонусов", "товары бонусом",
        )
    }
    matrix: dict[str, str] = {}
    for h, idx in hm.items():
        if h in known:
            continue
        val = _raw_cell_by_index(row, idx)
        original_header = _raw_cell_by_index(header_row, idx)
        matrix[(original_header or h)] = val
    if not matrix:
        return ""
    return json.dumps(matrix, ensure_ascii=False)


# Обратное отображение: русское название → ключ БД
_CLIENT_TYPE_REVERSE: dict[str, str] = {
    v.lower(): k for k, v in CLIENT_TYPES.items()
}
# Дополнительные алиасы для устойчивости
_CLIENT_TYPE_REVERSE.update({
    "торговая сеть": "retail_chain",
    "дистрибьютор": "distributor",
    "дистрибьютор": "distributor",
    "оптовик": "wholesaler",
    "обычный": "regular",
    "обычный клиент": "regular",
    "retail_chain": "retail_chain",
    "distributor": "distributor",
    "wholesaler": "wholesaler",
    "regular": "regular",
})


def _parse_client_type(raw: str) -> str:
    return _CLIENT_TYPE_REVERSE.get(raw.strip().lower(), "regular")


# ─────────────────────────────────────────────────────────────
# ИМПОРТ КЛИЕНТОВ
# ─────────────────────────────────────────────────────────────

def import_clients(conn: sqlite3.Connection, path: Path) -> ImportReport:
    rep = ImportReport()
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if not header:
        rep.errors.append("Файл клиентов пуст.")
        return rep
    hm = _header_map(header)
    for line_no, row in enumerate(rows, start=2):
        if row is None or _row_is_empty(row):
            continue
        ext              = _cell(row, hm, "id клиента", "id")
        name             = _cell(row, hm, "название", "наименование")
        inn              = normalize_inn(_cell(row, hm, "инн"))
        contacts         = _cell(row, hm, "контакты")
        addresses        = _cell(row, hm, "адреса", "адрес")
        unload           = _cell(row, hm, "пункты разгрузки", "пункт разгрузки")
        contact_person   = _cell(row, hm, "контактное лицо")
        email            = _cell(row, hm, "электронная почта", "email", "e-mail")
        city_region_zip  = _cell(row, hm, "город/штат/почтовый индекс", "город")
        c_name           = _cell(row, hm, "название компании грузополучателя")
        c_contact        = _cell(row, hm, "контактное лицо грузополучателя")
        c_address        = _cell(row, hm, "адрес грузополучателя")
        c_city           = _cell(row, hm, "город/штат/почтовый индекс грузополучателя")
        c_phone          = _cell(row, hm, "телефон грузополучателя")
        c_email          = _cell(row, hm, "электронная почта грузополучателя",
                                  "email грузополучателя", "e-mail грузополучателя")
        # ← новое поле
        client_type_raw  = _cell(row, hm, "тип клиента", "категория клиента", "тип")
        client_type      = _parse_client_type(client_type_raw)

        if not name and not inn and not ext:
            continue
        try:
            if ext:
                existing = conn.execute(
                    "SELECT id FROM clients WHERE external_id = ?", (ext,)
                ).fetchone()
            else:
                existing = None
            kwargs = dict(
                external_id=ext or None, name=name, inn=inn,
                contacts=contacts, addresses=addresses, unload_points=unload,
                contact_person=contact_person, email=email,
                city_region_zip=city_region_zip,
                consignee_name=c_name, consignee_contact_person=c_contact,
                consignee_address=c_address, consignee_city_region_zip=c_city,
                consignee_phone=c_phone, consignee_email=c_email,
                client_type=client_type,
            )
            if existing:
                clients.update(conn, int(existing[0]), is_new=False, **kwargs)
            else:
                clients.insert(conn, is_new=False, **kwargs)
            rep.clients_rows += 1
        except Exception as e:  # noqa: BLE001
            rep.errors.append(f"Клиенты, строка {line_no}: {e}")
    audit.log(conn, "import", "clients", str(path))
    return rep


# ─────────────────────────────────────────────────────────────
# ИМПОРТ ТОВАРОВ
# ─────────────────────────────────────────────────────────────

def import_products(conn: sqlite3.Connection, path: Path) -> ImportReport:
    rep = ImportReport()
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if not header:
        rep.errors.append("Файл товаров пуст.")
        return rep
    hm = _header_map(header)
    for line_no, row in enumerate(rows, start=2):
        if row is None or _row_is_empty(row):
            continue
        ext               = _cell(row, hm, "id товара", "id")
        name              = _cell(row, hm, "наименование", "название")
        price             = _float_cell(row, hm, "базовая цена", "цена")
        box_barcode       = _cell(row, hm, "баркод коробки", "баркод")
        unit              = _cell(row, hm, "ед. измерения", "ед измерения", "ед.")
        units_per_box     = _float_cell(row, hm, "количество в кор. (шт)", "количество в кор", "шт в коробе")
        regular_piece     = _float_cell(row, hm, "цена за штуку (руб), регулярная цена",
                                         "цена за штуку", "регулярная цена штуки")
        boxes_per_pallet  = _float_cell(row, hm, "количество на паллете (кор)", "коробов на паллете")
        gross_weight      = _float_cell(row, hm, "масса брутто, (кг)", "масса брутто")
        volume            = _float_cell(row, hm, "объем, (м3)", "объём, (м3)", "объем м3", "объём м3")
        # ← новые логистические поля
        boxes_in_row      = _int_cell(row, hm, "коробов в ряде", "кол-во коробов в ряде", "boxes_in_row")
        rows_per_pallet   = _int_cell(row, hm, "рядов в паллете", "кол-во рядов в паллете", "rows_per_pallet")
        pallet_height_mm  = _int_cell(row, hm, "высота с паллетой (мм)", "высота паллеты мм", "pallet_height_mm")
        box_dimensions    = _cell(row, hm, "размер короба д*ш*в", "размер короба", "box_dimensions")

        if not name and ext == "":
            continue
        if price is None:
            rep.errors.append(f"Товары, строка {line_no}: нет базовой цены.")
            continue
        try:
            p = products.by_external_id(conn, ext) if ext else None
            common = dict(
                box_barcode=box_barcode,
                unit=unit or "кор",
                units_per_box=int(units_per_box or 0),
                regular_piece_price=float(regular_piece or 0),
                boxes_per_pallet=float(boxes_per_pallet or 0),
                gross_weight_kg=float(gross_weight or 0),
                volume_m3=float(volume or 0),
                boxes_in_row=int(boxes_in_row or 0),
                rows_per_pallet=int(rows_per_pallet or 0),
                pallet_height_mm=int(pallet_height_mm or 0),
                box_dimensions=box_dimensions,
            )
            if p:
                products.update(conn, p.id, external_id=ext or None,
                                name=name or p.name, base_price=price,
                                box_barcode=common["box_barcode"] or p.box_barcode,
                                unit=common["unit"] or p.unit,
                                units_per_box=common["units_per_box"] or p.units_per_box,
                                regular_piece_price=common["regular_piece_price"] or p.regular_piece_price,
                                boxes_per_pallet=common["boxes_per_pallet"] or p.boxes_per_pallet,
                                gross_weight_kg=common["gross_weight_kg"] or p.gross_weight_kg,
                                volume_m3=common["volume_m3"] or p.volume_m3,
                                boxes_in_row=common["boxes_in_row"] or p.boxes_in_row,
                                rows_per_pallet=common["rows_per_pallet"] or p.rows_per_pallet,
                                pallet_height_mm=common["pallet_height_mm"] or p.pallet_height_mm,
                                box_dimensions=common["box_dimensions"] or p.box_dimensions)
            else:
                products.insert(conn, external_id=ext or None,
                                name=name or "", base_price=price, **common)
            rep.products_rows += 1
        except Exception as e:  # noqa: BLE001
            rep.errors.append(f"Товары, строка {line_no}: {e}")
    audit.log(conn, "import", "products", str(path))
    return rep


# ─────────────────────────────────────────────────────────────
# ИМПОРТ АКЦИЙ (без изменений)
# ─────────────────────────────────────────────────────────────

def import_promotions(conn: sqlite3.Connection, path: Path) -> ImportReport:
    rep = ImportReport()
    required_headers = ("id товара", "размер скидки")
    try:
        rows_data = _load_rows_with_openpyxl(path)
    except Exception as e_openpyxl:
        try:
            rows_data = _load_rows_with_calamine(path)
        except Exception as e_calamine:
            rep.errors.append(
                f"Не удалось прочитать файл акций. openpyxl: {e_openpyxl}. calamine: {e_calamine}"
            )
            return rep
    if not rows_data:
        rep.errors.append("Файл акций пуст.")
        return rep
    header_info = _find_header_row(rows_data, required_headers)
    if header_info is None:
        rep.errors.append(
            "Не найдена строка заголовков акций (ожидаются колонки 'ID товара' и 'Размер скидки')."
        )
        return rep
    header_idx, header_row, hm = header_info
    seen_products: set[str] = set()
    for line_no, row in enumerate(rows_data[header_idx + 1:], start=header_idx + 2):
        if row is None or _row_is_empty(row):
            continue
        ext       = _cell(row, hm, "id товара", "id")
        ptype     = _cell(row, hm, "тип акции", "тип")
        disc      = _float_cell(row, hm, "размер скидки", "скидка", "скидка %")
        d1s       = _cell(row, hm, "дата начала", "начало")
        d2s       = _cell(row, hm, "дата окончания", "окончание", "конец")
        bonus_raw = _cell(row, hm,
            "id товаров-бонусов (через запятую)", "id товаров бонуса",
            "id товаров-бонусов", "бонусные id", "id бонусных товаров",
            "id бонусов", "товары бонусом",
        )
        if not ext:
            rep.errors.append(f"Акции, строка {line_no}: нет ID товара.")
            continue
        if ext in seen_products:
            rep.errors.append(f"Акции, строка {line_no}: дубль ID товара {ext}.")
            continue
        seen_products.add(ext)
        if disc is None:
            rep.errors.append(f"Акции, строка {line_no}: нет размера скидки.")
            continue
        try:
            d1 = parse_dmY(d1s)
            d2 = parse_dmY(d2s)
        except ValueError as e:
            rep.errors.append(f"Акции, строка {line_no}: {e}")
            continue
        if d1 > d2:
            rep.errors.append(f"Акции, строка {line_no}: дата начала позже окончания.")
            continue
        pr = products.by_external_id(conn, ext)
        if not pr:
            rep.errors.append(f"Акции, строка {line_no}: товар «{ext}» не найден.")
            continue
        parsed_bonus = parse_product_external_ids_csv(bonus_raw)
        bonus_norm   = normalize_product_external_ids_csv(bonus_raw)
        if parsed_bonus:
            miss = missing_product_external_ids(conn, parsed_bonus)
            if miss:
                rep.errors.append(
                    f"Акции, строка {line_no}: не найдены товары-бонусы: {', '.join(miss)}."
                )
                continue
        try:
            promotions.upsert(conn, pr.id,
                promo_type=ptype, discount_percent=float(disc),
                valid_from_iso=iso(d1), valid_to_iso=iso(d2),
                bonus_other_product_ids=bonus_norm,
                matrix_rules_json=_extract_matrix_rules_json(row, header_row, hm),
            )
            rep.promotions_rows += 1
        except Exception as e:
            rep.errors.append(f"Акции, строка {line_no}: {e}")
    audit.log(conn, "import", "promotions", str(path))
    return rep


# ─────────────────────────────────────────────────────────────
# ЭКСПОРТ КЛИЕНТОВ
# ─────────────────────────────────────────────────────────────

def export_clients(conn: sqlite3.Connection, path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "clients"
    ws.append([
        "ID клиента", "Название", "ИНН", "Контакты", "Адреса",
        "Пункты разгрузки", "Контактное лицо", "Электронная почта",
        "Город/Штат/Почтовый индекс",
        "Название компании грузополучателя", "Контактное лицо грузополучателя",
        "Адрес грузополучателя", "Город/Штат/Почтовый индекс грузополучателя",
        "Телефон грузополучателя", "Электронная почта грузополучателя",
        "Новый (0/1)",
        "Тип клиента",   # ← новое
    ])
    for c in clients.list_all(conn):
        ws.append([
            c.external_id or "", c.name, c.inn, c.contacts, c.addresses,
            c.unload_points, c.contact_person, c.email, c.city_region_zip,
            c.consignee_name, c.consignee_contact_person, c.consignee_address,
            c.consignee_city_region_zip, c.consignee_phone, c.consignee_email,
            1 if c.is_new else 0,
            c.client_type_label,   # ← читаемое название: «Дистрибутор» и т.д.
        ])
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    audit.log(conn, "export", "clients", str(path))


# ─────────────────────────────────────────────────────────────
# ЭКСПОРТ ТОВАРОВ
# ─────────────────────────────────────────────────────────────

def export_products(conn: sqlite3.Connection, path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "products"
    ws.append([
        "ID товара", "Наименование", "Базовая цена", "Баркод коробки",
        "Ед. измерения", "Количество в кор. (ШТ)",
        "Цена за штуку (РУБ), регулярная цена",
        "Количество на паллете (КОР)", "масса брутто, (КГ)", "объём, (м3)",
        "Коробов в ряде",          # ← новое
        "Рядов в паллете",         # ← новое
        "Высота с паллетой (мм)",  # ← новое
        "Размер короба д*ш*в",     # ← новое
    ])
    for p in products.list_all(conn):
        ws.append([
            p.external_id or "", p.name, p.base_price, p.box_barcode,
            p.unit, p.units_per_box, p.regular_piece_price,
            p.boxes_per_pallet, p.gross_weight_kg, p.volume_m3,
            p.boxes_in_row, p.rows_per_pallet,
            p.pallet_height_mm, p.box_dimensions,
        ])
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    audit.log(conn, "export", "products", str(path))


# ─────────────────────────────────────────────────────────────
# ЭКСПОРТ АКЦИЙ (с matrix_rules)
# ─────────────────────────────────────────────────────────────

# Фиксированный порядок известных ключей matrix_rules для читаемого экспорта
_MATRIX_KEYS_ORDER = [
    "prepay_25", "prepay_50",
    "volume_300", "volume_500",
    "expiry_pct", "expiry_rub",
    "promo_15_2_qty", "promo_10_3_qty",
    "promo_date_from", "promo_date_to",
]


def export_promotions(conn: sqlite3.Connection, path: Path) -> None:
    all_promos = promotions.list_all(conn)

    # Собираем все уникальные ключи matrix_rules из всех записей
    extra_keys: list[str] = list(_MATRIX_KEYS_ORDER)
    seen_extra: set[str] = set(_MATRIX_KEYS_ORDER)
    for r in all_promos:
        if r.matrix_rules_json:
            try:
                mr = json.loads(r.matrix_rules_json)
                for k in mr:
                    if k not in seen_extra:
                        extra_keys.append(k)
                        seen_extra.add(k)
            except (json.JSONDecodeError, TypeError):
                pass

    wb = Workbook()
    ws = wb.active
    ws.title = "promotions"
    header = [
        "ID товара", "Тип акции", "Размер скидки",
        "Дата начала", "Дата окончания",
        "ID товаров-бонусов (через запятую)",
    ] + extra_keys
    ws.append(header)

    for r in all_promos:
        d1 = parse_iso(r.valid_from_iso) if r.valid_from_iso else None
        d2 = parse_iso(r.valid_to_iso)   if r.valid_to_iso   else None
        mr: dict[str, Any] = {}
        if r.matrix_rules_json:
            try:
                mr = json.loads(r.matrix_rules_json)
            except (json.JSONDecodeError, TypeError):
                mr = {}
        row = [
            r.product_external_id or "", r.promo_type, r.discount_percent,
            format_dmY(d1) if d1 else "", format_dmY(d2) if d2 else "",
            r.bonus_other_product_ids or "",
        ] + [mr.get(k, "") for k in extra_keys]
        ws.append(row)

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    audit.log(conn, "export", "promotions", str(path))