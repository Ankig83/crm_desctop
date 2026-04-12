from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────────────────────
# Dataclass
# ─────────────────────────────────────────────────────────────

@dataclass
class RusLine:
    external_id: str = ""
    box_barcode: str = ""
    name: str = ""
    unit: str = "кор"
    qty: float = 0.0
    base_price: float = 0.0
    regular_price_per_box: float = 0.0
    regular_price_per_piece: float = 0.0
    discount_percent: float = 0.0
    line_total: float = 0.0
    units_per_box: int = 0
    boxes_per_pallet: int = 0
    gross_weight_kg: float = 0.0
    volume_m3: float = 0.0
    boxes_in_row: int = 0
    rows_per_pallet: int = 0
    pallet_height_mm: int = 0
    box_dimensions: str = ""
    is_bonus: bool = False
    matrix_rules: dict[str, Any] = field(default_factory=dict)


# ─────────────────────────────────────────────────────────────
# Стили
# ─────────────────────────────────────────────────────────────

FONT_NAME = "Arial"
_THIN   = Side(style="thin")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_C_HEADER_BG  = "1F3864"
_C_HEADER_FG  = "FFFFFF"
_C_LABEL_BG   = "D9E1F2"
_C_VALUE_BG   = "FFFFFF"
_C_TYPE_BG    = "E2EFDA"   # светло-зелёный — тип клиента (теперь в H3)
_C_TYPE_FG    = "1F4E79"
_C_COL_HDR    = "BDD7EE"
_C_TOTAL_BG   = "1F4E79"
_C_TOTAL_FG   = "FFFFFF"
_C_DATA_ALT   = "F2F2F2"
_C_BONUS_BG   = "FFF2CC"   # светло-жёлтый — бонусные строки


def _f(bold=False, size=9, color="000000") -> Font:
    return Font(name=FONT_NAME, bold=bold, size=size, color=color)

def _fill(rgb: str) -> PatternFill:
    return PatternFill("solid", fgColor=rgb)

def _al(h="left", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _set(cell, value=None, *, bold=False, size=9, fg="000000",
         bg=None, h="left", wrap=False, border=False) -> None:
    if value is not None:
        cell.value = value
    cell.font = _f(bold=bold, size=size, color=fg)
    if bg:
        cell.fill = _fill(bg)
    cell.alignment = _al(h=h, wrap=wrap)
    if border:
        cell.border = _BORDER


# ─────────────────────────────────────────────────────────────
# Хелперы для matrix_rules
# ─────────────────────────────────────────────────────────────

def _mr_prepay_disc(mr: dict) -> float:
    """Лучшая скидка за предоплату из matrix_rules, %."""
    best = 0.0
    for key, val in mr.items():
        if key.startswith("prepay_"):
            try:
                best = max(best, float(val or 0))
            except (ValueError, TypeError):
                pass
    return best


def _mr_volume_disc(mr: dict, threshold: int) -> float:
    """Скидка за объём для конкретного порога (300 или 500), %."""
    try:
        return float(mr.get(f"volume_{threshold}") or 0)
    except (ValueError, TypeError):
        return 0.0


def _mr_promo_rules(mr: dict) -> list[tuple[float, int]]:
    """[(threshold_qty, same_qty), ...] — акционные правила, сортировка по threshold."""
    names = sorted(
        key[len("promo_"):-len("_qty")]
        for key in mr
        if key.startswith("promo_") and key.endswith("_qty")
    )
    rules = []
    for name in names:
        try:
            threshold = float(name.split("_")[0])
            same_qty = int(float(mr.get(f"promo_{name}_qty", 0) or 0))
            if threshold > 0:
                rules.append((threshold, same_qty))
        except (ValueError, IndexError):
            pass
    return rules


# ─────────────────────────────────────────────────────────────
# Структура строк (СООТВЕТСТВУЕТ ШАБЛОНУ заказчика)
# ─────────────────────────────────────────────────────────────
# R1   No: {order_no}
# R2   Информация о Покупателе*
# R3   Название Компании          [H = "Тип клиента: XXX  Скидка: -N%"]
# R4   ИНН
# R5   Контактное Лицо
# R6   Адрес
# R7   Город/Штат/Почтовый Индекс
# R8   Телефон
# R9   Электронная почта
# R10  Информация о Грузополучателе*
# R11  Название Компании (груз.)
# R12  Контактное Лицо (груз.)
# R13  Адрес (груз.)
# R14  Город/Штат/... (груз.)
# R15  Телефон (груз.)
# R16  Электронная почта (груз.)
# R17  ДАТА ЗАКАЗА               ← критично для 1С
# R18  ДАТА ДОСТАВКИ             ← критично для 1С
# R19  подсказки
# R20  заголовки колонок
# R21  нумерация
# R22+ данные товаров

_DATA_START = 22


def export_rus_variant_a(
    path: Path,
    *,
    client,
    quote_date: date,
    lines: list[RusLine],
    delivery_date: date | None = None,
    order_no: str = "",
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "RUS"

    # ── Ширины колонок ────────────────────────────────────────
    col_widths = {
        1: 12, 2: 18, 3: 42, 4: 7, 5: 9, 6: 12, 7: 14,
        8: 12, 9: 11, 10: 11, 11: 11, 12: 11, 13: 11,
        14: 10, 15: 13, 16: 13, 17: 13, 18: 13,
        19: 11, 20: 11, 21: 9, 22: 12, 23: 10,
        24: 16, 25: 16, 26: 16,
        27: 10, 28: 10, 29: 13, 30: 5, 31: 14,
    }
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    # ── Строка 1: No ──────────────────────────────────────────
    ws.merge_cells("B1:T1")
    no_text = f"No: {order_no}" if order_no else "No: …......"
    _set(ws["B1"], no_text, bold=True, size=11, fg=_C_HEADER_FG, bg=_C_HEADER_BG)
    ws.row_dimensions[1].height = 18

    # ── Строка 2: Информация о Покупателе ─────────────────────
    ws.merge_cells("B2:G2")
    _set(ws["B2"], "Информация о Покупателе*", bold=True, size=10,
         fg=_C_HEADER_FG, bg=_C_HEADER_BG)
    ws.row_dimensions[2].height = 16

    # ── Строки 3–9: данные покупателя ────────────────────────
    buyer_rows = [
        (3,  "Название Компании",          client.name if client else ""),
        (4,  "ИНН",                        client.inn  if client else ""),
        (5,  "Контактное Лицо",            client.contact_person if client else ""),
        (6,  "Адрес",                      client.addresses if client else ""),
        (7,  "Город/Штат/Почтовый Индекс", client.city_region_zip if client else ""),
        (8,  "Телефон",                    client.contacts if client else ""),
        (9,  "Электронная почта",          client.email if client else ""),
    ]
    for row_num, label, value in buyer_rows:
        ws.row_dimensions[row_num].height = 15
        lc = ws.cell(row=row_num, column=2, value=label)
        _set(lc, bold=True, bg=_C_LABEL_BG, border=True)
        vc = ws.cell(row=row_num, column=3, value=value)
        ws.merge_cells(start_row=row_num, start_column=3,
                       end_row=row_num, end_column=7)
        _set(vc, bg=_C_VALUE_BG, border=True)

    # ── Тип клиента — в ячейку H3 (как "Базовый прайс-лист" в шаблоне) ──
    # Не добавляем отдельную строку, чтобы не сдвигать нумерацию для 1С
    client_type_label = ""
    if client:
        if hasattr(client, "client_type_label"):
            client_type_label = client.client_type_label
        elif hasattr(client, "client_type"):
            from crm_desktop.repositories.clients import CLIENT_TYPES
            client_type_label = CLIENT_TYPES.get(client.client_type, client.client_type)
    if client_type_label:
        disc_suffix = ""
        if client and hasattr(client, "type_discount_pct") and client.type_discount_pct > 0:
            disc_suffix = f"  (скидка -{client.type_discount_pct:.0f}%)"
        h3 = ws.cell(row=3, column=8,
                     value=f"Тип клиента: {client_type_label}{disc_suffix}")
        h3.font = _f(bold=True, size=9, color=_C_TYPE_FG)
        h3.fill = _fill(_C_TYPE_BG)
        h3.alignment = _al(h="left")

    # ── Строка 10: Информация о Грузополучателе ───────────────
    ws.merge_cells("B10:G10")
    _set(ws["B10"], "Информация о Грузополучателе*", bold=True, size=10,
         fg=_C_HEADER_FG, bg=_C_HEADER_BG)
    ws.row_dimensions[10].height = 16

    # ── Строки 11–16: данные грузополучателя ──────────────────
    cons_rows = [
        (11, "Название Компании",          client.consignee_name if client else ""),
        (12, "Контактное Лицо",            client.consignee_contact_person if client else ""),
        (13, "Адрес",                      client.consignee_address if client else ""),
        (14, "Город/Штат/Почтовый Индекс", client.consignee_city_region_zip if client else ""),
        (15, "Телефон",                    client.consignee_phone if client else ""),
        (16, "Электронная почта",          client.consignee_email if client else ""),
    ]
    for row_num, label, value in cons_rows:
        ws.row_dimensions[row_num].height = 15
        lc = ws.cell(row=row_num, column=2, value=label)
        _set(lc, bold=True, bg=_C_LABEL_BG, border=True)
        vc = ws.cell(row=row_num, column=3, value=value)
        ws.merge_cells(start_row=row_num, start_column=3,
                       end_row=row_num, end_column=7)
        _set(vc, bg=_C_VALUE_BG, border=True)

    # ── Строка 17: ДАТА ЗАКАЗА (критично для 1С) ─────────────
    ws.row_dimensions[17].height = 15
    _set(ws.cell(row=17, column=2), "ДАТА ЗАКАЗА",
         bold=True, bg=_C_LABEL_BG, border=True)
    dvc = ws.cell(row=17, column=3, value=quote_date.strftime("%d.%m.%Y"))
    ws.merge_cells(start_row=17, start_column=3, end_row=17, end_column=7)
    _set(dvc, bg=_C_VALUE_BG, border=True, h="center")

    # ── Строка 18: ДАТА ДОСТАВКИ ──────────────────────────────
    ws.row_dimensions[18].height = 15
    _set(ws.cell(row=18, column=2), "ДАТА ДОСТАВКИ",
         bold=True, bg=_C_LABEL_BG, border=True)
    ddvc = ws.cell(row=18, column=3,
                   value=delivery_date.strftime("%d.%m.%Y") if delivery_date else "")
    ws.merge_cells(start_row=18, start_column=3, end_row=18, end_column=7)
    _set(ddvc, bg=_C_VALUE_BG, border=True, h="center")

    # ── Строка 19: подсказки ──────────────────────────────────
    ws.row_dimensions[19].height = 13
    ws.merge_cells("B19:C19")
    ws["E19"].value = "Кол-во"
    ws["F19"].value = "Цена"
    ws["G19"].value = "Итоговая цена"
    ws.merge_cells("AA19:AC19")
    ws["AA19"].value = "Размеры паллеты (1200*1000) финский и кол-во"
    for ref in ("E19", "F19", "G19", "AA19"):
        _set(ws[ref], size=8, fg="595959", h="center")

    # ── Строка 20: заголовки колонок ─────────────────────────
    ws.row_dimensions[20].height = 32
    col_headers = {
        1:  "Артикул",
        2:  "Баркод коробки",
        3:  "Наименование Товара",
        4:  "Ед. Измерения",
        5:  "Заказ (КОР)",
        6:  "Цена итог (КОР)",
        7:  "Итоговая цена заказа",
        8:  "Цена за Штуку (РУБ), регулярная цена",
        9:  "Предоплата -2%",
        10: "Объем > 300 кор",
        11: "Объем > 500 кор",
        12: "Акция 15+2 (КОФЕ)",
        13: "Акция 10+3 (КОНФ)",
        14: "Количество в кор. (ШТ)",
        15: "Цена Короба рег. прайс (РУБ)",
        16: "Цена Короба -2% (РУБ)",
        17: "Цена Короба -6% (РУБ)",
        18: "Цена Короба -8% (РУБ)",
        19: "Доп. скидка (РУБ)*",
        20: "Количество на паллете (КОР)",
        21: "Итого Паллет",
        22: "масса брутто, (КГ)",
        23: "объём, (м3)",
        24: "Регулярная Итоговая сумма заказа (РУБ)",
        25: "Итоговая сумма заказа + акция (РУБ)",
        26: "Итоговая сумма заказа, акция +мот (РУБ)",
        27: "кол-во коробов в ряде",
        28: "кол-во рядов в паллете",
        29: "высота с паллетой (мм)",
        31: "размер короба д*ш*в",
    }
    for col, label in col_headers.items():
        c = ws.cell(row=20, column=col, value=label)
        c.font = _f(bold=True, size=8)
        c.fill = _fill(_C_COL_HDR)
        c.border = _BORDER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ── Строка 21: нумерация ──────────────────────────────────
    ws.row_dimensions[21].height = 12
    num_map = {2: 1, 3: 2, 4: 3, 8: 4, 14: 5, 15: 6, 16: 7,
               19: 8, 20: 9, 21: 11, 22: 12, 23: 14, 24: 13,
               25: 13, 26: 13, 27: 16, 28: 17, 29: 18, 31: 20}
    for col, num in num_map.items():
        c = ws.cell(row=21, column=col, value=num)
        c.font = _f(size=7, color="595959")
        c.alignment = _al(h="center")

    # ── Строки данных (с 22) ──────────────────────────────────
    last_data_row = _DATA_START + len(lines) - 1

    for idx, line in enumerate(lines):
        r = _DATA_START + idx
        ws.row_dimensions[r].height = 15
        mr = line.matrix_rules

        if line.is_bonus:
            _write_bonus_row(ws, r, line)
            continue

        price_box   = line.regular_price_per_box or line.base_price
        price_piece = line.regular_price_per_piece or (
            price_box / line.units_per_box if line.units_per_box else 0.0
        )
        # Итоговая цена за коробку с учётом скидки
        price_after = price_box * (1 - line.discount_percent / 100) if line.discount_percent else price_box
        # Итого строки = qty × price_after
        line_sum    = line.qty * price_after if line.qty else 0.0
        # Цена за штуку после скидки
        price_piece_after = price_after / line.units_per_box if line.units_per_box else price_after

        pallets    = line.qty / line.boxes_per_pallet if line.boxes_per_pallet and line.qty else 0.0
        mass_total = line.gross_weight_kg * line.qty if line.gross_weight_kg and line.qty else 0.0
        vol_total  = line.volume_m3 * line.qty if line.volume_m3 and line.qty else 0.0

        # Регулярная цена короба = price_box × units_per_box (до скидки)
        price_reg = round(price_box * line.units_per_box, 4) if line.units_per_box else round(price_box, 4)

        # ── Коэффициенты скидок из matrix_rules ──────────────
        prepay_disc = _mr_prepay_disc(mr) or 2.0       # дефолт 2% если не задан
        vol_300_disc = _mr_volume_disc(mr, 300) or 6.0  # дефолт 6%
        vol_500_disc = _mr_volume_disc(mr, 500) or 8.0  # дефолт 8%

        # C9-C11: коэффициент (1 = нет скидки, 0.98 = -2%)
        prepay_coeff = round(1.0 - prepay_disc / 100, 6) if mr.get("prepay_25") or any(
            k.startswith("prepay_") for k in mr) else 1.0
        vol_300_coeff = round(1.0 - vol_300_disc / 100, 6) if mr.get("volume_300") or any(
            k == "volume_300" for k in mr) else 1.0
        vol_500_coeff = round(1.0 - vol_500_disc / 100, 6) if mr.get("volume_500") or any(
            k == "volume_500" for k in mr) else 1.0

        # C12: коэффициент первой акции (threshold/(threshold+same_qty))
        # C13: кол-во коробок второй акции
        promo_rules = _mr_promo_rules(mr)
        promo1_coeff = 1.0
        promo2_count = 0
        if promo_rules:
            thr, same = promo_rules[0]
            if thr > 0 and same > 0:
                promo1_coeff = round(thr / (thr + same), 6)
        if len(promo_rules) > 1:
            _, same2 = promo_rules[1]
            promo2_count = same2

        # C16-C18: цена за короб при каждой скидке (эталонные значения)
        price_m_prepay = round(price_reg * (1 - prepay_disc / 100), 2)
        price_m_v300   = round(price_reg * (1 - vol_300_disc / 100), 2)
        price_m_v500   = round(price_reg * (1 - vol_500_disc / 100), 2)

        # C19: доп. скидка в рублях (expiry_rub из matrix_rules или расчётная)
        try:
            _erub = float(mr.get("expiry_rub") or 0)
        except (ValueError, TypeError):
            _erub = 0.0
        if _erub > 0:
            extra_disc = _erub
        else:
            extra_disc = round(price_reg - price_after * line.units_per_box, 2) if line.units_per_box else 0.0

        data = {
            1:  line.external_id,
            2:  line.box_barcode,
            3:  line.name,
            4:  line.unit,
            5:  line.qty or None,
            # C6 = итого строки (qty × price_after) — соответствует шаблону
            6:  round(line_sum, 2) or None,
            # C7 = цена за штуку после скидки — соответствует шаблону
            7:  round(price_piece_after, 4) or None,
            8:  round(price_piece, 4) or None,
            9:  prepay_coeff,
            10: vol_300_coeff,
            11: vol_500_coeff,
            12: promo1_coeff,
            13: promo2_count or None,
            14: line.units_per_box or None,
            15: round(price_reg, 4) or None,
            16: price_m_prepay or None,
            17: price_m_v300 or None,
            18: price_m_v500 or None,
            19: extra_disc if extra_disc > 0 else None,
            20: line.boxes_per_pallet or None,
            21: round(pallets, 2) or None,
            22: round(mass_total, 3) or None,
            23: round(vol_total, 4) or None,
            24: round(line.qty * price_box, 2) if line.qty and price_box else None,
            25: round(line_sum, 2) or None,
            26: None,
            27: line.boxes_in_row or None,
            28: line.rows_per_pallet or None,
            29: line.pallet_height_mm or None,
            31: line.box_dimensions or None,
        }

        bg = _C_VALUE_BG if idx % 2 == 0 else _C_DATA_ALT

        for col, val in data.items():
            c = ws.cell(row=r, column=col, value=val)
            c.border = _BORDER
            c.fill   = _fill(bg)
            c.font   = _f(size=9)
            if col == 3:
                c.alignment = _al(h="left")
            elif col in (24, 25, 26):
                c.font = _f(size=9, bold=True)
                c.alignment = _al(h="right")
            else:
                c.alignment = _al(h="center")

    # ── Строка ИТОГО ─────────────────────────────────────────
    regular_rows = [
        _DATA_START + i
        for i, ln in enumerate(lines)
        if not ln.is_bonus
    ]

    total_row = last_data_row + 1
    ws.row_dimensions[total_row].height = 16
    ws.merge_cells(start_row=total_row, start_column=1,
                   end_row=total_row, end_column=4)
    tc = ws.cell(row=total_row, column=1, value="ИТОГО регулярный ассортимент")
    _set(tc, bold=True, size=9, fg=_C_TOTAL_FG, bg=_C_TOTAL_BG, h="right", border=True)

    for col in range(1, 32):
        if col <= 4:
            continue
        c = ws.cell(row=total_row, column=col)
        c.fill   = _fill(_C_TOTAL_BG)
        c.font   = _f(bold=True, size=9, color=_C_TOTAL_FG)
        c.border = _BORDER
        c.alignment = _al(h="right")

    if regular_rows:
        def _sum_formula(col_letter: str) -> str:
            parts = [f"{col_letter}{r}" for r in regular_rows]
            return "=" + "+".join(parts)

        for col_num in (5, 6, 7, 21, 22, 23, 24, 25, 26):
            cl = get_column_letter(col_num)
            ws.cell(row=total_row, column=col_num).value = _sum_formula(cl)

    # ── Заморозка ─────────────────────────────────────────────
    ws.freeze_panes = f"A{_DATA_START}"

    # ── Сохранение ───────────────────────────────────────────
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _write_bonus_row(ws, r: int, line: RusLine) -> None:
    """Бонусная строка: товар бесплатно — цена и суммы = 0."""
    data = {
        1:  line.external_id,
        2:  line.box_barcode,
        3:  f"БОНУС: {line.name}",
        4:  line.unit,
        5:  line.qty or None,
        6:  0,
        7:  0,
        8:  None,
        9:  None, 10: None, 11: None, 12: None, 13: None,
        14: line.units_per_box or None,
        15: None, 16: None, 17: None, 18: None, 19: None,
        20: line.boxes_per_pallet or None,
        21: None, 22: None, 23: None,
        24: 0,
        25: 0,
        26: None,
        27: None, 28: None, 29: None, 31: None,
    }
    for col, val in data.items():
        c = ws.cell(row=r, column=col, value=val)
        c.border = _BORDER
        c.fill   = _fill(_C_BONUS_BG)
        c.font   = _f(size=9, bold=(col == 3), color="1F4E79")
        if col == 3:
            c.alignment = _al(h="left")
        else:
            c.alignment = _al(h="center")
