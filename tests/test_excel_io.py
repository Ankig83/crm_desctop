"""Тесты импорта и экспорта данных через Excel."""
from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from crm_desktop.adapters.excel_io import (
    export_clients,
    export_products,
    export_promotions,
    import_clients,
    import_products,
    import_promotions,
)
from crm_desktop.repositories import clients, products, promotions


# ─────────────────────────────────────────────────────────────────
# Хелперы
# ─────────────────────────────────────────────────────────────────

def _clients_wb(rows: list[tuple]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.append(("ID клиента", "Название", "ИНН", "Контакты", "Адреса", "Пункты разгрузки"))
    for r in rows:
        ws.append(r)
    return wb


def _products_wb(rows: list[tuple]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.append(("ID товара", "Наименование", "Базовая цена"))
    for r in rows:
        ws.append(r)
    return wb


def _promotions_wb(rows: list[tuple]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.append(("ID товара", "Тип акции", "Размер скидки", "Дата начала", "Дата окончания"))
    for r in rows:
        ws.append(r)
    return wb


def _save(wb: Workbook, path: Path) -> Path:
    wb.save(path)
    return path


# ─────────────────────────────────────────────────────────────────
# Импорт клиентов
# ─────────────────────────────────────────────────────────────────

class TestImportClients:
    def test_basic_import(self, conn, tmp_path: Path):
        p = _save(_clients_wb([("C-001", "ООО Тест", "7712345678", "", "", "")]),
                  tmp_path / "c.xlsx")
        rep = import_clients(conn, p)
        assert rep.errors == []
        lst = clients.list_all(conn)
        assert len(lst) == 1
        assert lst[0].name == "ООО Тест"
        assert lst[0].inn == "7712345678"

    def test_import_two_clients(self, conn, tmp_path: Path):
        p = _save(_clients_wb([
            ("C-001", "Клиент А", "1111111111", "", "", ""),
            ("C-002", "Клиент Б", "2222222222", "", "", ""),
        ]), tmp_path / "c.xlsx")
        rep = import_clients(conn, p)
        assert rep.errors == []
        assert len(clients.list_all(conn)) == 2

    def test_reimport_updates_existing(self, conn, tmp_path: Path):
        p = tmp_path / "c.xlsx"
        _save(_clients_wb([("C-001", "Старое", "3333333333", "", "", "")]), p)
        import_clients(conn, p)

        _save(_clients_wb([("C-001", "Новое", "3333333333", "", "", "")]), p)
        import_clients(conn, p)
        lst = clients.list_all(conn)
        assert len(lst) == 1
        assert lst[0].name == "Новое"

    def test_missing_header_skips_all_rows(self, conn, tmp_path: Path):
        """Файл без нужных колонок — строки пропускаются без ошибок (не падает)."""
        wb = Workbook()
        ws = wb.active
        ws.append(("Только одна колонка",))
        ws.append(("значение",))
        p = _save(wb, tmp_path / "bad.xlsx")
        rep = import_clients(conn, p)
        # Строки не импортируются (нет совпадающих заголовков)
        assert rep.clients_rows == 0

    def test_empty_rows_skipped(self, conn, tmp_path: Path):
        wb = Workbook()
        ws = wb.active
        ws.append(("ID клиента", "Название", "ИНН", "Контакты", "Адреса", "Пункты разгрузки"))
        ws.append((None, None, None, None, None, None))  # пустая строка
        p = _save(wb, tmp_path / "empty_rows.xlsx")
        rep = import_clients(conn, p)
        assert rep.clients_rows == 0


# ─────────────────────────────────────────────────────────────────
# Импорт товаров
# ─────────────────────────────────────────────────────────────────

class TestImportProducts:
    def test_basic_import(self, conn, tmp_path: Path):
        p = _save(_products_wb([("P-001", "Кофе Гранд", 250.0)]), tmp_path / "p.xlsx")
        rep = import_products(conn, p)
        assert rep.errors == []
        lst = products.list_all(conn)
        assert len(lst) == 1
        assert lst[0].external_id == "P-001"
        assert lst[0].base_price == pytest.approx(250.0)

    def test_update_on_reimport(self, conn, tmp_path: Path):
        p = tmp_path / "p.xlsx"
        _save(_products_wb([("P-001", "Кофе", 100.0)]), p)
        import_products(conn, p)

        _save(_products_wb([("P-001", "Кофе Премиум", 200.0)]), p)
        import_products(conn, p)
        lst = products.list_all(conn)
        assert len(lst) == 1
        assert lst[0].name == "Кофе Премиум"
        assert lst[0].base_price == pytest.approx(200.0)

    def test_zero_price_allowed(self, conn, tmp_path: Path):
        p = _save(_products_wb([("P-FREE", "Бонус", 0)]), tmp_path / "p.xlsx")
        rep = import_products(conn, p)
        assert rep.errors == []

    def test_multiple_products(self, conn, tmp_path: Path):
        p = _save(_products_wb([
            ("A", "Товар А", 10.0),
            ("B", "Товар Б", 20.0),
            ("C", "Товар В", 30.0),
        ]), tmp_path / "p.xlsx")
        import_products(conn, p)
        assert len(products.list_all(conn)) == 3

    def test_missing_header_skips_rows(self, conn, tmp_path: Path):
        """Файл без нужных колонок — строки пропускаются."""
        wb = Workbook()
        ws = wb.active
        ws.append(("Заголовок",))
        ws.append(("данные",))
        p = _save(wb, tmp_path / "bad.xlsx")
        rep = import_products(conn, p)
        assert rep.products_rows == 0


# ─────────────────────────────────────────────────────────────────
# Импорт акций
# ─────────────────────────────────────────────────────────────────

class TestImportPromotions:
    def test_basic_import(self, conn, tmp_path: Path):
        products.insert(conn, external_id="P-001", name="Товар", base_price=100.0)
        p = _save(_promotions_wb([("P-001", "Сезон", 10, "01.04.2026", "30.04.2026")]),
                  tmp_path / "pr.xlsx")
        rep = import_promotions(conn, p)
        assert rep.errors == []
        pid = products.by_external_id(conn, "P-001").id
        row = promotions.get_for_product(conn, pid)
        assert row.discount_percent == pytest.approx(10.0)

    def test_invalid_product_id_error(self, conn, tmp_path: Path):
        """Акция на несуществующий товар → ошибка."""
        p = _save(_promotions_wb([("NOEXIST", "Сезон", 5, "01.04.2026", "30.04.2026")]),
                  tmp_path / "pr.xlsx")
        rep = import_promotions(conn, p)
        assert len(rep.errors) > 0

    def test_invalid_date_error(self, conn, tmp_path: Path):
        products.insert(conn, external_id="P-001", name="Товар", base_price=100.0)
        p = _save(_promotions_wb([("P-001", "Сезон", 5, "не дата", "тоже не дата")]),
                  tmp_path / "pr.xlsx")
        rep = import_promotions(conn, p)
        assert len(rep.errors) > 0

    def test_import_with_bonus_rules_column(self, conn, tmp_path: Path):
        """Колонка bonus_rules должна сохраниться в matrix_rules_json."""
        products.insert(conn, external_id="P-001", name="Товар", base_price=100.0)
        bonus_json = json.dumps([
            {"threshold": 10, "same_qty": 1, "fixed_id": "", "fixed_qty": 1, "choice_ids": ""}
        ])
        wb = Workbook()
        ws = wb.active
        ws.append(("ID товара", "Тип акции", "Размер скидки", "Дата начала", "Дата окончания", "bonus_rules"))
        ws.append(("P-001", "Сезон", 5, "01.04.2026", "30.04.2026", bonus_json))
        p = _save(wb, tmp_path / "pr.xlsx")
        rep = import_promotions(conn, p)
        assert rep.errors == []
        pid = products.by_external_id(conn, "P-001").id
        row = promotions.get_for_product(conn, pid)
        assert "bonus_rules" in row.matrix_rules_json

    def test_duplicate_product_id_in_file_error(self, conn, tmp_path: Path):
        """Два одинаковых ID товара в файле → ошибка дубликата."""
        products.insert(conn, external_id="P-DUP", name="Товар", base_price=100.0)
        p = _save(_promotions_wb([
            ("P-DUP", "Сезон", 5, "01.04.2026", "30.04.2026"),
            ("P-DUP", "Зима",  8, "01.11.2026", "28.02.2027"),
        ]), tmp_path / "pr.xlsx")
        rep = import_promotions(conn, p)
        assert len(rep.errors) > 0


# ─────────────────────────────────────────────────────────────────
# Экспорт и round-trip
# ─────────────────────────────────────────────────────────────────

class TestExportClients:
    def test_export_creates_file(self, conn, tmp_path: Path):
        clients.insert(conn,
                       external_id="EXP-001", name="Тест Экспорт", inn="9999999999",
                       contacts="", addresses="", unload_points="",
                       client_type="distributor")
        out = tmp_path / "exp_clients.xlsx"
        export_clients(conn, out)
        assert out.exists()
        wb = load_workbook(out)
        ws = wb.active
        all_values = [ws.cell(row=r, column=c).value
                      for r in range(1, ws.max_row + 1)
                      for c in range(1, ws.max_column + 1)]
        assert "Тест Экспорт" in all_values

    def test_export_empty_db(self, conn, tmp_path: Path):
        out = tmp_path / "empty.xlsx"
        export_clients(conn, out)
        assert out.exists()


class TestExportProducts:
    def test_export_contains_data(self, conn, tmp_path: Path):
        products.insert(conn, external_id="EXP-P1", name="Товар экспорт", base_price=300.0)
        out = tmp_path / "exp_products.xlsx"
        export_products(conn, out)
        wb = load_workbook(out)
        ws = wb.active
        all_values = [ws.cell(row=r, column=c).value
                      for r in range(1, ws.max_row + 1)
                      for c in range(1, ws.max_column + 1)]
        assert "EXP-P1" in all_values or "Товар экспорт" in all_values


class TestExportPromotions:
    def test_export_includes_bonus_rules_column(self, conn, tmp_path: Path):
        pid = products.insert(conn, external_id="P-BR", name="Товар с бонусом", base_price=100.0)
        bonus_rules = json.dumps([
            {"threshold": 10, "same_qty": 1, "fixed_id": "", "fixed_qty": 1, "choice_ids": ""}
        ])
        promotions.upsert(conn, pid, promo_type="", discount_percent=5,
                          valid_from_iso="2026-01-01", valid_to_iso="2026-12-31",
                          matrix_rules_json=json.dumps({"bonus_rules": bonus_rules}))
        out = tmp_path / "exp_promos.xlsx"
        export_promotions(conn, out)
        wb = load_workbook(out)
        ws = wb.active
        headers = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
        assert "bonus_rules" in headers

    def test_export_then_reimport_round_trip(self, conn, tmp_path: Path):
        """Экспорт → импорт сохраняет discount_percent."""
        pid = products.insert(conn, external_id="RT-001", name="Round-trip", base_price=150.0)
        promotions.upsert(conn, pid, promo_type="Сезон", discount_percent=12.0,
                          valid_from_iso="2026-04-01", valid_to_iso="2026-09-30")
        out = tmp_path / "rt_promos.xlsx"
        export_promotions(conn, out)

        promotions.delete_for_product(conn, pid)
        rep = import_promotions(conn, out)
        assert rep.errors == []
        row = promotions.get_for_product(conn, pid)
        assert row is not None
        assert row.discount_percent == pytest.approx(12.0)
