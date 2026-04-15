"""Тесты экспорта в формат RUS.xlsx."""
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from crm_desktop.adapters.rus_export import RusLine, export_rus_variant_a

_HEADER_ROW = 20  # _DATA_START(22) - 2
_DATA_ROW_1 = 22  # _DATA_START


def _export(tmp_path: Path, lines: list[RusLine]) -> Path:
    out = tmp_path / "RUS.xlsx"
    export_rus_variant_a(
        out,
        client=None,
        quote_date=date(2026, 4, 7),
        lines=lines,
    )
    return out


def _headers(ws, row: int) -> list:
    return [ws.cell(row=row, column=i).value for i in range(1, ws.max_column + 1)]


# ─────────────────────────────────────────────────────────────────
# Фиксированные заголовки строки 20
# ─────────────────────────────────────────────────────────────────

def test_fixed_column_headers_present(tmp_path: Path) -> None:
    """Строка 20 содержит ожидаемые фиксированные заголовки."""
    out = _export(tmp_path, [
        RusLine(
            external_id="P1", box_barcode="", name="Товар 1", unit="кор",
            qty=1, regular_price_per_box=100, regular_price_per_piece=10,
            units_per_box=10, boxes_per_pallet=20,
            gross_weight_kg=1, volume_m3=0.1,
            base_price=100, discount_percent=0, line_total=100,
            matrix_rules={},
        ),
    ])
    wb = load_workbook(out, data_only=True)
    ws = wb.active
    headers = _headers(ws, _HEADER_ROW)
    assert "Артикул" in headers
    assert "Наименование Товара" in headers
    assert "Предоплата -2%" in headers


def test_two_lines_data_rows(tmp_path: Path) -> None:
    """Два товара → данные в строках 22 и 23."""
    out = _export(tmp_path, [
        RusLine(
            external_id="P1", box_barcode="", name="Товар 1", unit="кор",
            qty=5, regular_price_per_box=100, regular_price_per_piece=10,
            units_per_box=10, boxes_per_pallet=20,
            gross_weight_kg=1, volume_m3=0.1,
            base_price=100, discount_percent=0, line_total=500,
            matrix_rules={},
        ),
        RusLine(
            external_id="P2", box_barcode="", name="Товар 2", unit="кор",
            qty=3, regular_price_per_box=200, regular_price_per_piece=20,
            units_per_box=10, boxes_per_pallet=20,
            gross_weight_kg=2, volume_m3=0.2,
            base_price=200, discount_percent=0, line_total=600,
            matrix_rules={},
        ),
    ])
    wb = load_workbook(out, data_only=True)
    ws = wb.active
    # Артикул = колонка 1
    assert ws.cell(row=_DATA_ROW_1,     column=1).value == "P1"
    assert ws.cell(row=_DATA_ROW_1 + 1, column=1).value == "P2"


def test_order_number_in_b1(tmp_path: Path) -> None:
    """Номер заказа записывается в ячейку B1."""
    out = tmp_path / "RUS_no.xlsx"
    export_rus_variant_a(
        out,
        client=None,
        quote_date=date(2026, 4, 7),
        lines=[],
        order_no="42",
    )
    wb = load_workbook(out, data_only=True)
    ws = wb.active
    b1 = str(ws["B1"].value or "")
    assert "42" in b1


def test_export_empty_lines_no_exception(tmp_path: Path) -> None:
    """Экспорт без строк товаров не должен падать."""
    out = tmp_path / "empty_rus.xlsx"
    export_rus_variant_a(out, client=None, quote_date=date(2026, 4, 7), lines=[])
    assert out.exists()
